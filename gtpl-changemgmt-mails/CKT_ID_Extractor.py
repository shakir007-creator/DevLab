import win32com.client
import os
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Alignment
from pathlib import Path

def extract_data():

    # Access Outlook
    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        namespace = outlook.GetNamespace("MAPI")
        # print("\nSuccessfully connected to Outlook.")
    except Exception as e:
        print(f"\nError accessing Outlook: {e}")
        raise

    # Access the specific folder: support.ahmedabad@gtpl.net > Inbox > Changemgmt_Airtel
    try:
        folder = namespace.Folders["My Outlook Data File(1)"].Folders["Inbox"].Folders["Changemgmt_Airtel"]
        # print("\nSuccessfully accessed folder: Changemgmt_Airtel.")
    except Exception as e:
        print(f"\nError accessing folder: {e}")
        raise

    # Get all emails in the folder
    try:
        messages = folder.Items
        # print(f"Found {len(messages)} emails in the folder.")
    except Exception as e:
        print(f"\nError retrieving emails: {e}")
        raise

    # Function to extract a section of text between two delimiters
    def extract_section(body, start_delim, end_delim):
        try:
            start_index = body.find(start_delim)
            if start_index == -1:
                return None
            start_index += len(start_delim)
            end_index = body.find(end_delim, start_index)
            if end_index == -1:
                end_index = len(body)
            return body[start_index:end_index].strip()
        except Exception as e:
            print(f"\nError extracting section from body: {e}")
            return None

    # Function to extract CKT IDs from the email body using the extract_section function
    def extract_ckts_from_body(email_body):
        try:
            # Extract the section containing CKT IDs
            identity_table = extract_section(email_body, "Impacted LSI : Party Name", "We request you")
            if not identity_table:
                return []

            identity_lines = identity_table.split('\n')
            ckt_ids = []
            for line in identity_lines:
                parts = line.split(":")
                if len(parts) != 2:
                    continue
                identity_number = parts[0].strip()

                # Check if the identity number starts with "1Old_" and remove it if present
                if identity_number.startswith("1Old_"):
                    identity_number = identity_number[5:]  # Remove "1Old_" prefix

                ckt_ids.append(int(identity_number))
            return ckt_ids
        except Exception as e:
            print(f"\nError extracting CKT IDs: {e}")
            return []

    # Path to the existing Excel file
    file_path = Path.cwd() / "Excel_File" / "Latest_B2B.xlsx"

    # Prepare a list to hold the CKT data
    ckt_data = []

    processed_emails = 0
    skipped_emails = 0

    # Process each email
    for message in messages:
        try:
            if message.Unread:  # Only process unread emails
                mail_body = message.Body
                start_greet = mail_body.index("With respect")
                end_greet = mail_body.index("Details of Maintenance")
                greet = mail_body[start_greet:end_greet-1]
                start_sub = message.Subject.index('CRQ: “') + len('CRQ: “')
                end_sub = message.Subject.index('” Planned Maintenance activity')
                crq = message.Subject[start_sub:end_sub]
                full_date1 = message.Subject.index("scheduled on ") + len("scheduled on ")
                full_date2 = message.Subject.index('-202')
                full_date = message.Subject[full_date1:full_date2+5].strip()
                activity_date = message.Subject[full_date1:full_date2].strip()

                # Skip emails with "Postponed" or "Completed successfully" in the subject or body
                if "Postponed" in greet or "Completed successfully" in greet:
                    print(f"\nSkipped Email --> CRQ: {crq}. Might be Postponed or Completed.\n")
                    print("============================================================================\n")
                    skipped_emails += 1
                    message.UnRead = True
                    message.Save()
                    continue

                # Extract CKT IDs from the email body
                ckt_ids = extract_ckts_from_body(mail_body)

                if ckt_ids:
                    # Add the extracted CKT IDs to the list
                    ckt_data.extend(ckt_ids)

                # If the file exists, append data to the sheet
                if os.path.exists(file_path):
                    try:
                        wb = load_workbook(file_path)
                        sheet = wb['Lookup-Sheet']  # Access the existing sheet
                    except InvalidFileException as e:
                        print(f"\nError loading workbook: {e}")
                        continue
                    except Exception as e:
                        print(f"\nUnexpected error loading workbook: {e}")
                        continue

                    # Check if the CRQ number is already present in column D (without "CRQ:" prefix)
                    crq_exists = False
                    for row in range(1, sheet.max_row + 1):
                        if sheet[f'D{row}'].value:  # Check if cell in column D has a value
                            existing_value = str(sheet[f'D{row}'].value)  # Convert to string to avoid any issues
                            # Compare only the CRQ number (find if crq exists anywhere in the existing cell value)
                            if str(crq) in existing_value:  # Check if the CRQ number is present in the cell
                                crq_exists = True
                                break

                    if crq_exists:
                        print(f"\nCRQ: {crq} already exists. Skipping this email.")
                        print("\n+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++\n")
                        skipped_emails += 1
                        message.UnRead = True
                        message.Save()
                        ckt_data.clear()
                        continue  # Skip to the next email

                    # Find the last non-empty row in column D
                    last_row = 1
                    for row in range(1, sheet.max_row + 1):
                        if sheet[f'D{row}'].value:
                            last_row = row

                    # Calculate the starting row for the new data
                    start_row = last_row + 2  # Leave 2 empty rows before writing new data

                    # Write the CRQ and activity date to the next empty row in column D
                    sheet[f'D{start_row}'] = f"CRQ: {crq} ({activity_date})"
                    sheet[f'D{start_row}'].alignment = Alignment(horizontal='center', vertical='center')

                    # Write the CKT IDs into Column D while keeping other formatting intact
                    for idx, value in enumerate(ckt_data, start=start_row + 1):
                        # Write the CKT ID to Column D
                        sheet[f'D{idx}'].value = value  # Write the value without changing formatting
                        sheet[f'D{idx}'].alignment = Alignment(horizontal='center', vertical='center')

                # Mark the email as read
                message.UnRead = False
                ckt_data.clear()
                processed_emails += 1
                print(f'\nE-Mail for CRQ: {crq}, Dated: {full_date} processed successfully.\n')
                print("################################")

                # Save the workbook with formatting preserved (no changes to formatting)
                try:
                    wb.save(file_path)
                # print(f"\nData for CRQ {crq} saved successfully.")
                except Exception as e:
                    print(f"\nError saving workbook: {e}")
                    continue

        except Exception as e:
            print(f"\nError processing email with subject {message.Subject}: {e}")
            continue

    # print(f"\nData written and saved to the Excel file: {file_path}\n")
    if processed_emails != 0:
        print(f"\n\nData Extracted from {processed_emails} Emails successfully.\n")

    if skipped_emails != 0:
        print(f"\n{skipped_emails} Emails Skipped.")

    print("\n======================================================================\n\n")


if __name__ == "__main__":
    extract_data()
    
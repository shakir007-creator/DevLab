from openpyxl import load_workbook
from pathlib import Path
from openpyxl.styles import Alignment


def insert_formula():

    file_path = Path.cwd() / "Excel_File" / "Latest_B2B.xlsx"

    print("\nInserting Formula in the sheet...\n")

    try:
        # Load the workbook and sheet
        wb = load_workbook(file_path)
        sheet = wb['Lookup-Sheet']  # Access the sheet by name

        # List of columns to insert formulas
        columns_to_insert = ['A', 'B', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']

        # Find the last non-empty row in column D (starting from the bottom)
        last_row = sheet.max_row
        while last_row > 1 and not sheet[f'D{last_row}'].value:
            last_row -= 1

        # Formula you provided
        formula = '''=IF(OR(@'Lookup-Sheet'!$D:D = "", ISNUMBER(SEARCH("CRQ",@D:D ))), "",IFERROR(IFERROR(INDEX(B2B!$1:$1048576,MATCH(@'Lookup-Sheet'!$D:$D,B2B!$AD:$AD,0),MATCH(@'Lookup-Sheet'!$A$1:$M$1,B2B!$A$1:$AK$1,0)),INDEX('Link Shifting'!$1:$1048576,MATCH(@'Lookup-Sheet'!$D:$D,'Link Shifting'!$AC:$AC,0),MATCH(@'Lookup-Sheet'!$A$1:$M$1,'Link Shifting'!$A$1:$AD$1,0))), "No Data Found"))'''

        # Formula for column C (new formula you provided)
        formula_c = '''=IF(@D:D="","",IF(ISNUMBER(SEARCH("CRQ",@D:D)),D:D,C{previous_row}))'''

        # Insert the formula into the specified columns up to the last non-empty row in column D
        for row in range(2, last_row + 1):  # Start from row 2 to last_row
            # Insert the formula in the specified columns for the current row
            for col in columns_to_insert:
                try:
                    cell = sheet[f'{col}{row}']  # Reference the cell
                    cell.value = formula  # Insert the formula
                    cell.alignment = Alignment(horizontal='center', vertical='center')  # Apply centering alignment
                except Exception as e:
                    print(f"\nError inserting formula in column {col}, row {row}: {e}\n")

            # Insert the formula for column C with dynamic reference to the previous row
            try:
                cell_c = sheet[f'C{row}']  # Reference the cell in column C
                dynamic_formula_c = formula_c.format(previous_row=row-1)  # Replace {previous_row}
                cell_c.value = dynamic_formula_c  # Insert the formula in column C
                cell_c.alignment = Alignment(horizontal='center', vertical='center')  # Apply centering alignment
            except Exception as e:
                print(f"\nError inserting formula in column C, row {row}: {e}\n")

        # Save the changes to the workbook
        try:
            wb.save(file_path)
            print("\nInserted Formula Successfully.\n")
        except Exception as e:
            print(f"\nError saving the workbook: {e}\n")

    except FileNotFoundError as e:
        print(f"\nError: The file {file_path} was not found. Please check the file path.\n")
    except KeyError as e:
        print(f"\nError: The sheet 'Lookup-Sheet' was not found in the workbook.\n")
    except Exception as e:
        print(f"\nAn unexpected error occurred: {e}\n")

    print("\n======================================================================\n\n")


if __name__ == "__main__":
    insert_formula()

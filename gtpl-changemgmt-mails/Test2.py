# import win32com.client
# import os
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.styles import Alignment

# # Access Outlook
# outlook = win32com.client.Dispatch("Outlook.Application")
# namespace = outlook.GetNamespace("MAPI")
    
# # Access the specific folder: support.ahmedabad@gtpl.net > Inbox > Changemgmt_Airtel
# folder = namespace.Folders["support.ahmedabad@gtpl.net"].Folders["Inbox"].Folders["Changemgmt_Airtel"]

# # Get all emails in the folder
# messages = folder.Items

# # Function to extract a section of text between two delimiters
# def extract_section(body, start_delim, end_delim):
#     start_index = body.find(start_delim)
#     if start_index == -1:
#         return None
#     start_index += len(start_delim)
#     end_index = body.find(end_delim, start_index)
#     if end_index == -1:
#         end_index = len(body)
#     return body[start_index:end_index].strip()

# # Function to extract CKT IDs from the email body using the extract_section function
# def extract_ckts_from_body(email_body):
#     # Extract the section containing CKT IDs
#     identity_table = extract_section(email_body, "Impacted LSI : Party Name", "We request you")
#     if not identity_table:
#         return []

#     identity_lines = identity_table.split('\n')
#     ckt_ids = []
#     for line in identity_lines:
#         parts = line.split(":")
#         if len(parts) != 2:
#             continue
#         identity_number = parts[0].strip()
#         ckt_ids.append(int(identity_number))
#     return ckt_ids

# # Path to the existing Excel file
# file_path = r'D:\Shared\Changemgmt\Excel_File\Test.xlsx'

# # Prepare a list to hold the CKT data
# ckt_data = []

# # Process each email
# for message in messages:
#     if message.Unread:  # Only process unread emails
#         mail_body = message.Body
#         start_greet = mail_body.index("With respect")
#         end_greet = mail_body.index("Details of Maintenance")
#         greet = mail_body[start_greet:end_greet-1]
#         start_sub = message.Subject.index('CRQ: “') + len('CRQ: “')
#         end_sub = message.Subject.index('” Planned Maintenance activity')
#         crq = message.Subject[start_sub:end_sub]
#         date1 = message.Subject.index("scheduled on ") + len("scheduled on ")
#         date2 = message.Subject.index('-202')
#         activity_date = message.Subject[date1:date2].strip()
        
#         # Skip emails with "Postponed" or "Completed successfully" in the subject or body
#         if "Postponed" in greet or "Completed successfully" in greet:
#             print(f"\nSkipped --> {crq}\n")
#             print("======================================\n")
#             message.UnRead = True
#             message.Save()
#             continue

#         # Extract CKT IDs from the email body
#         ckt_ids = extract_ckts_from_body(mail_body)
        
#         if ckt_ids:
#             # Add the extracted CKT IDs to the list
#             ckt_data.extend(ckt_ids)

#         # Create a DataFrame to write to Excel
#         df = pd.DataFrame(ckt_data, columns=["CKT_IDs"])  # Set column name as "CKT_IDs"

#         # If the file exists, append data to the sheet; otherwise, create a new file
#         if os.path.exists(file_path):
#             # Load workbook and select the sheet
#             wb = load_workbook(file_path)
#             sheet = wb['Sheet']  # Access the existing sheet
            
#             max_row = sheet.max_row  # Get the max row to append the data
#             start_row = max_row + 3  # Start appending from the next row
            
#             # Write the CRQ and activity date to the next empty row in column C
#             sheet[f'C{start_row}'] = f"CRQ: {crq} ({activity_date})"
            
#             # Write CKT data to column C, starting from the row after the header (start_row + 1)
#             for idx, value in enumerate(ckt_data, start=start_row + 1):
#                 sheet[f'C{idx}'] = value  # Write data to column 'C' (CKT_IDs)
            
#             # Apply formatting (center alignment) to the newly written cells
#             for idx in range(start_row, start_row + len(ckt_data) + 1):
#                 sheet[f'C{idx}'].alignment = Alignment(horizontal='center', vertical='center')

#             # Save the workbook
#             wb.save(file_path)
            
#         else:
#             # If the file does not exist, create it and add data to 'Sheet1'
#             df.to_excel(file_path, index=False, sheet_name='Sheet')

#         # Mark the email as read
#         message.UnRead = True
#         ckt_data.clear()

# print(f"Data written and saved to the Excel file: {file_path}")


# import win32com.client
# import os
# import pandas as pd
# from openpyxl import load_workbook

# # Access Outlook
# outlook = win32com.client.Dispatch("Outlook.Application")
# namespace = outlook.GetNamespace("MAPI")

# # Access the specific folder: support.ahmedabad@gtpl.net > Inbox > Changemgmt_Airtel
# folder = namespace.Folders["support.ahmedabad@gtpl.net"].Folders["Inbox"].Folders["Changemgmt_Airtel"]

# # Get all emails in the folder
# messages = folder.Items

# # Function to extract a section of text between two delimiters
# def extract_section(body, start_delim, end_delim):
#     start_index = body.find(start_delim)
#     if start_index == -1:
#         return None
#     start_index += len(start_delim)
#     end_index = body.find(end_delim, start_index)
#     if end_index == -1:
#         end_index = len(body)
#     return body[start_index:end_index].strip()

# # Function to extract CKT IDs from the email body using the extract_section function
# def extract_ckts_from_body(email_body):
#     # Extract the section containing CKT IDs
#     identity_table = extract_section(email_body, "Impacted LSI : Party Name", "We request you")
#     if not identity_table:
#         return []

#     identity_lines = identity_table.split('\n')
#     ckt_ids = []
#     for line in identity_lines:
#         parts = line.split(":")
#         if len(parts) != 2:
#             continue
#         identity_number = parts[0].strip()
#         ckt_ids.append(int(identity_number))
#     return ckt_ids

# # Path to the existing Excel file
# file_path = r'D:\Shared\Changemgmt\Excel_File\Test.xlsx'

# # Prepare a list to hold the CKT data
# ckt_data = []

# # Process each email
# for message in messages:
#     if message.Unread:  # Only process unread emails
#         mail_body = message.Body
#         start_greet = mail_body.index("With respect")
#         end_greet = mail_body.index("Details of Maintenance")
#         greet = mail_body[start_greet:end_greet-1]
#         start_sub = message.Subject.index('CRQ: “') + len('CRQ: “')
#         end_sub = message.Subject.index('” Planned Maintenance activity')
#         crq = message.Subject[start_sub:end_sub]
#         date1 = message.Subject.index("scheduled on ") + len("scheduled on ")
#         date2 = message.Subject.index('-202')
#         activity_date = message.Subject[date1:date2].strip()

#         # Skip emails with "Postponed" or "Completed successfully" in the subject or body
#         if "Postponed" in greet or "Completed successfully" in greet:
#             print(f"\nSkipped --> {crq}\n")
#             print("======================================\n")
#             message.UnRead = True
#             message.Save()
#             continue

#         # Extract CKT IDs from the email body
#         ckt_ids = extract_ckts_from_body(mail_body)

#         if ckt_ids:
#             # Add the extracted CKT IDs to the list
#             ckt_data.extend(ckt_ids)

#         # Create a DataFrame to write to Excel
#         df = pd.DataFrame(ckt_data, columns=["CKT_IDs"])  # Set column name as "CKT_IDs"

#         # If the file exists, append data to the sheet; otherwise, create a new file
#         if os.path.exists(file_path):
#             # Load workbook and select the sheet
#             wb = load_workbook(file_path)
#             sheet = wb['Sheet']  # Access the existing sheet
            
#             # Find the last non-empty row in column C
#             last_row = 1
#             for row in range(1, sheet.max_row + 1):
#                 if sheet[f'C{row}'].value:
#                     last_row = row

#             # Calculate the starting row for the new data
#             start_row = last_row + 3  # Leave 2 empty rows before writing new data

#             # Write the CRQ and activity date to the next empty row in column C
#             sheet[f'C{start_row}'] = f"CRQ: {crq} ({activity_date})"

#             # Write CKT data to column C, starting from the row after the header (start_row + 1)
#             for idx, value in enumerate(ckt_data, start=start_row + 1):
#                 sheet[f'C{idx}'] = value  # Write data to column 'C' (CKT_IDs)

#             # Save the workbook
#             wb.save(file_path)

#         else:
#             # If the file does not exist, create it and add data to 'Sheet1'
#             df.to_excel(file_path, index=False, sheet_name='Sheet')

#         # Mark the email as read
#         message.UnRead = True
#         ckt_data.clear()

# print(f"Data written and saved to the Excel file: {file_path}")



# import win32com.client
# import os
# import pandas as pd

# # Access Outlook
# outlook = win32com.client.Dispatch("Outlook.Application")
# namespace = outlook.GetNamespace("MAPI")
    
# # Access the specific folder: support.ahmedabad@gtpl.net > Inbox > Changemgmt_Airtel
# folder = namespace.Folders["support.ahmedabad@gtpl.net"].Folders["Inbox"].Folders["Changemgmt_Airtel"]

# # Get all emails in the folder
# messages = folder.Items

# # Function to extract a section of text between two delimiters
# def extract_section(body, start_delim, end_delim):
#     start_index = body.find(start_delim)
#     if start_index == -1:
#         return None
#     start_index += len(start_delim)
#     end_index = body.find(end_delim, start_index)
#     if end_index == -1:
#         end_index = len(body)
#     return body[start_index:end_index].strip()

# # Function to extract CKT IDs from the email body using the extract_section function
# def extract_ckts_from_body(email_body):
#     # Extract the section containing CKT IDs
#     identity_table = extract_section(email_body, "Impacted LSI : Party Name", "We request you")
#     if not identity_table:
#         return []

#     identity_lines = identity_table.split('\n')
#     ckt_ids = []
#     for line in identity_lines:
#         parts = line.split(":")
#         if len(parts) != 2:
#             continue
#         identity_number = parts[0].strip()
#         ckt_ids.append(int(identity_number))
#     return ckt_ids

# # Path to the existing Excel file
# file_path = r'D:\Shared\Changemgmt\Excel_File\Latest_B2B.xlsx'

# # Prepare a list to hold the CKT data
# ckt_data = []

# # Process each email
# for message in messages:
#     if message.Unread:  # Only process unread emails
#         mail_body = message.Body
#         start_greet = mail_body.index("With respect")
#         end_greet = mail_body.index("Details of Maintenance")
#         greet = mail_body[start_greet:end_greet-1]
#         start_sub = message.Subject.index('CRQ: “') + len('CRQ: “')
#         end_sub = message.Subject.index('” Planned Maintenance activity')
#         crq = message.Subject[start_sub:end_sub]
#         date1 = message.Subject.index("scheduled on ") + len("scheduled on ")
#         date2 = message.Subject.index('-202')
#         activity_date = message.Subject[date1:date2].strip()
        
#         # Skip emails with "Postponed" or "Completed successfully" in the subject or body
#         if "Postponed" in greet or "Completed successfully" in greet:
#             print(f"\nSkipped --> {crq}\n")
#             print("======================================\n")
#             message.UnRead = True
#             message.Save()
#             continue

#         # Extract CKT IDs from the email body
#         ckt_ids = extract_ckts_from_body(mail_body)
        
#         if ckt_ids:
#             # Add the extracted CKT IDs to the list
#             ckt_data.extend(ckt_ids)

#         # Create a DataFrame to write to Excel
#         df = pd.DataFrame(ckt_data, columns=["AIRTEL Circuit ID"])  # Set column name as "CKT_IDs"

#         # If the file exists, append data to the sheet; otherwise, create a new file
#         if os.path.exists(file_path):
#             with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
#                 # Check if 'Sheet' exists
#                 try:
#                     writer.sheets['Lookup-Sheet']
#                 except KeyError:
#                     print("Sheet Doesn't Exists")
#                 else:
#                     sheet = writer.sheets['Lookup-Sheet']  # Access the existing sheet
#                     # max_row = sheet.max_row  # Get the max row to append the data
#                     # start_row = max_row + 3  # Start appending from the next row
#                     # sheet[f'D{start_row}'].value = f"CRQ: {crq} ({activity_date})"

#                     # Find the last non-empty row in column C
#                     last_row = 1
#                     for row in range(1, sheet.max_row + 1):
#                         if sheet[f'D{row}'].value:
#                             last_row = row

#                     # Calculate the starting row for the new data
#                     start_row = last_row + 3  # Leave 2 empty rows before writing new data

#                     # Write the CRQ and activity date to the next empty row in column C
#                     sheet[f'D{start_row}'] = f"CRQ: {crq} ({activity_date})"

#                     for idx, value in enumerate(ckt_data, start=start_row):
#                         # Write the CKT ID to Column C
#                         sheet[f'D{idx+1}'].value = value  

#                         # Check Column A to see if it contains "Active" or "#N/A"
#                         cell_in_column_a = sheet[f'A{idx+1}'].value
#                         if cell_in_column_a == "Active":
#                             # If "Active", keep the CKT ID
#                             continue
#                         elif cell_in_column_a == "#N/A" or cell_in_column_a == None:
#                             # If #N/A error, remove the CKT ID and write the next one
#                             sheet[f'D{idx+1}'].value = None  # Remove CKT ID
#                             next_value = ckt_data[idx + 1] if idx + 1 < len(ckt_data) else None
#                             sheet[f'D{idx+1}'].value = next_value  # Write the next CKT ID
#                         else:
#                             # In case it's something else (such as an empty cell or unexpected value)
#                             continue
#         # else:
#         #     # If the file does not exist, create it and add data to 'Sheet'
#         #     df.to_excel(file_path, index=False, sheet_name='Sheet')

#         # Mark the email as read
#         message.UnRead = True
#         ckt_data.clear()

# print(f"Data written and saved to the Excel file: {file_path}")


import win32com.client
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# Access Outlook
outlook = win32com.client.Dispatch("Outlook.Application")
namespace = outlook.GetNamespace("MAPI")

# Access the specific folder: support.ahmedabad@gtpl.net > Inbox > Changemgmt_Airtel
folder = namespace.Folders["support.ahmedabad@gtpl.net"].Folders["Inbox"].Folders["Changemgmt_Airtel"]

# Get all emails in the folder
messages = folder.Items

# Function to extract a section of text between two delimiters
def extract_section(body, start_delim, end_delim):
    start_index = body.find(start_delim)
    if start_index == -1:
        return None
    start_index += len(start_delim)
    end_index = body.find(end_delim, start_index)
    if end_index == -1:
        end_index = len(body)
    return body[start_index:end_index].strip()

# Function to extract CKT IDs from the email body using the extract_section function
def extract_ckts_from_body(email_body):
    # Extract the section containing CKT IDs
    identity_table = extract_section(email_body, "Impacted LSI : Party Name", "We request you")
    if not identity_table:
        return []

    identity_lines = identity_table.split('\n')
    ckt_ids = []
    for line in identity_lines:
        parts = line.split(":")
        if len(parts) != 2:
            continue
        identity_number = parts[0].strip()
        ckt_ids.append(int(identity_number))
    return ckt_ids

# Path to the existing Excel file
file_path = r'D:\Shared\Changemgmt\Excel_File\Latest_B2B.xlsx'

# Prepare a list to hold the CKT data
ckt_data = []

# Process each email
for message in messages:
    if message.Unread:  # Only process unread emails
        mail_body = message.Body
        start_greet = mail_body.index("With respect")
        end_greet = mail_body.index("Details of Maintenance")
        greet = mail_body[start_greet:end_greet-1]
        start_sub = message.Subject.index('CRQ: “') + len('CRQ: “')
        end_sub = message.Subject.index('” Planned Maintenance activity')
        crq = message.Subject[start_sub:end_sub]
        date1 = message.Subject.index("scheduled on ") + len("scheduled on ")
        date2 = message.Subject.index('-202')
        activity_date = message.Subject[date1:date2].strip()

        # Skip emails with "Postponed" or "Completed successfully" in the subject or body
        if "Postponed" in greet or "Completed successfully" in greet:
            print(f"\nSkipped --> {crq}\n")
            print("======================================\n")
            message.UnRead = True
            message.Save()
            continue

        # Extract CKT IDs from the email body
        ckt_ids = extract_ckts_from_body(mail_body)

        if ckt_ids:
            # Add the extracted CKT IDs to the list
            ckt_data.extend(ckt_ids)

        # If the file exists, append data to the sheet
        if os.path.exists(file_path):
            # Open workbook with openpyxl to preserve formatting
            try:
                wb = load_workbook(file_path)
                sheet = wb['Lookup-Sheet']  # Access the existing sheet
            except InvalidFileException as e:
                print(f"Error loading workbook: {e}")
                continue

            # Find the last non-empty row in column D
            last_row = 1
            for row in range(1, sheet.max_row + 1):
                if sheet[f'D{row}'].value:
                    last_row = row

            # Calculate the starting row for the new data
            start_row = last_row + 3  # Leave 2 empty rows before writing new data

            # Write the CRQ and activity date to the next empty row in column D
            sheet[f'D{start_row}'] = f"CRQ: {crq} ({activity_date})"

            for idx, value in enumerate(ckt_data, start=start_row + 1):
                # Write the CKT ID to Column D
                sheet[f'D{idx}'] = value  

            # Save the workbook with formatting preserved
            wb.save(file_path)

        # Mark the email as read
        message.UnRead = True
        ckt_data.clear()

print(f"Data written and saved to the Excel file: {file_path}")




